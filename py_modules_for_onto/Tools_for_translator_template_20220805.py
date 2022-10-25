from py_modules_for_onto.Central_functions_for_Ontology_mod_20220810 import *
import pandas as pd
import copy
import os
import openpyxl
from openpyxl import load_workbook, Workbook
from collections import OrderedDict

FilePath = 'C:/Users/smhhborg/Documents/GitHub/Hendrik_Borgelt_Masterthesis/' \
           'Translator_test_modified_and_ready_for_testing_3.xlsx'

def start_rename_dict_chapters_for_translator_3(dict_int, list_of_keys_int):
    for i in list(dict_int):
        dict_int[i]['CFX Command Language for Run'] = rename_dict_chapters_for_translator_3(
            dict_int[i]['CFX Command Language for Run'], list_of_keys_int)
        if 'SOLUTION UNITS' in dict_int[i].keys():
            test_12345 = True
    return dict_int


# def create_merged_dict(dict_to_be_merged):
#     dict_merged = OrderedDict()
#     for i_4_int in dict_to_be_merged:
#         merge(dict_merged, dict_to_be_merged[i_4_int])
#     return dict_merged


def search_for_name_source_in_dict(dictionary):
    for k, v in list(dictionary.items()):
        source_name_changed = False
        if type(v) == dict:
            for k_2, v_2 in list(dictionary[k].items()):
                if '\\' in k_2:
                    test_1 = k + '\\' + '\\'.join(k_2.split('\\')[1:])
                    test_2 = test_1.split('=')
                    if len(test_2) == 2:
                        test_4 = copy.deepcopy(dictionary)
                        dictionary[test_2[0].rstrip(' ')] = test_2[1].lstrip(' ')
                        dictionary.pop(k)
                        # dictionary[k][k + '\\' + '\\'.join(k_2.split('\\')[1:])] = dictionary[k].pop(k_2)
                        source_name_changed = True
                elif '/' in k_2:
                    test_1 = k + '\\' + '\\'.join(k_2.split('\\')[1:])
                    test_2 = test_1.split('=')
                    if len(test_2) == 2:
                        test_4 = copy.deepcopy(dictionary)
                        dictionary[test_2[0].rstrip(' ')] = test_2[1].lstrip(' ')
                        dictionary.pop(k)
                        # dictionary[k][k + '\\' + '\\'.join(k_2.split('/')[1:])] = dictionary[k].pop(k_2)
                        source_name_changed = True
            if not source_name_changed:
                search_for_name_source_in_dict(dictionary[k])
            else:
                test = k
                test_5 = True

def search_for_name_option_in_dict(dictionary):
    for k, v in list(dictionary.items()):
        if type(v) == dict:
            for k_2, v_2 in list(dictionary[k].items()):
                if 'Option' in k_2:
                    dictionary[k][''.join(k) + ' ' + ''.join(k_2)] = dictionary[k].pop(k_2)
            search_for_name_option_in_dict(dictionary[k])


def start_rename_dict_chapters_for_translator_3(dict_int, list_of_keys_int):
    for i in list(dict_int):
        dict_int[i]['CFX Command Language for Run'] = rename_dict_chapters_for_translator_3(
            dict_int[i]['CFX Command Language for Run'], list_of_keys_int)
        if 'SOLUTION UNITS' in dict_int[i].keys():
            test_12345 = True
    return dict_int


def rename_dict_chapters_for_translator_3(dict_int, list_of_keys_int, current_name_int='', test_tuple=('', '', False)):
    for k_int, v_int in list(dict_int.items()):
        test_for_test_tuple = check_tuple_as_list_for_match_and_concatenate(k_int, list_of_keys_int, test_tuple)
        if k_int == 'SOLUTION UNITS':
            test = True
        if isinstance(v_int, str):
            for i_int in list_of_keys_int:
                if k_int == i_int:
                    pass
                else:
                    dict_int[current_name_int + ' ' + k_int] = v_int
            del dict_int[k_int]
        elif isinstance(v_int, (dict, OrderedDict)):
            if test_for_test_tuple[2]:
                dict_int[test_for_test_tuple[0]] = rename_dict_chapters_for_translator_3(dict_int[k_int],
                                                                                         list_of_keys_int,
                                                                                         test_for_test_tuple[0],
                                                                                         test_for_test_tuple)
                if test_for_test_tuple[0] != k_int:
                    dict_int.pop(k_int)
            else:
                dict_int[test_for_test_tuple[0] + ' ' + k_int] = rename_dict_chapters_for_translator_3(dict_int[k_int],
                                                                                                       list_of_keys_int,
                                                                                                       test_for_test_tuple[
                                                                                                           0],
                                                                                                       test_for_test_tuple)
                dict_int.pop(k_int)
        else:
            print('ERROR not dict/OrderedDict/str as value but type: ', type(v_int))
        if k_int == 'SOLUTION UNITS':
            test = True
    return dict_int


def check_tuple_as_list_for_match_and_concatenate(new_key_value, list_of_keys_int, list_to_be_checked):
    if new_key_value == 'ANALYSIS TYPE':
        test = True
    list_of_keys_int_2 = copy.deepcopy(list_of_keys_int)
    for i_1, i_2 in enumerate(list_of_keys_int_2):
        if isinstance(i_2, tuple):
            new_str_for_list = ''
            for i_3 in i_2:
                new_str_for_list += ' ' + i_3
            list_of_keys_int_2[i_1] = new_str_for_list.lstrip(' ')
    if (list_to_be_checked[0] + ' ' + new_key_value) in list_of_keys_int_2:
        list_after_check = (list_to_be_checked[0] + ' ' + new_key_value, '', True)
    elif (list_to_be_checked[1] + ' ' + new_key_value) in list_of_keys_int_2:
        list_after_check = (list_to_be_checked[1] + ' ' + new_key_value, '', True)
    elif new_key_value in list_of_keys_int_2:
        list_after_check = (new_key_value, '', True)
    else:
        list_after_check = (list_to_be_checked[0], new_key_value, False)
    if new_key_value == 'ANALYSIS TYPE':
        test_2 = True
    return list_after_check


def create_excel_file_for_translator(path_int, all_sims_int, list_of_keys_int):
    test_dict_for_translator_3 = copy.deepcopy(all_sims_int)
    search_for_name_source_in_dict(test_dict_for_translator_3)
    search_for_name_option_in_dict(test_dict_for_translator_3)
    list_of_keys_int_2 = copy.deepcopy(list_of_keys_int )
    test_dict_for_translator_4 = start_rename_dict_chapters_for_translator_3(test_dict_for_translator_3,
                                                                             list_of_keys_int_2)
    merged_dict_for_translator_2 = create_merged_dict(test_dict_for_translator_4)
    flattened_dict_2 = flatten_dict_to_list(merged_dict_for_translator_2['CFX Command Language for Run'])
    flattened_dict_2 = list(dict.fromkeys(flattened_dict_2))
    df_trans_2 = pd.DataFrame(flattened_dict_2)
    if not os.path.exists(path_int):
        ExcelWorkbook = Workbook()
    else:
        ExcelWorkbook = load_workbook(path_int)
    writer = pd.ExcelWriter(path_int, engine='openpyxl')
    writer.book = ExcelWorkbook
    df_trans_2.to_excel(writer, header=False, index=False, sheet_name='List_of_new_names_2')
    writer.save()
    writer.close()
    # flattened_dict = list(dict.fromkeys(flattened_dict))
    # df_trans = pd.DataFrame(flattened_dict)
    # pd.DataFrame(df_trans).to_excel('Translator_test.xlsx', header=False, index=False)
    # list_dict_names_2 = list(dict_of_excel_translator['Dict_name'].values())
    #
    # conv_dict = add_own_kpis_and_save_figs(all_sims)


def flatten_dict_to_list(dict_int):
    flattened_dict_list_int = []
    for key_int, value_int in dict_int.items():
        if key_int in flattened_dict_list_int:
            print(key_int)
        else:
            flattened_dict_list_int.append(key_int)
            if isinstance(value_int, (dict, OrderedDict)):
                if isinstance(flatten_dict_to_list(dict_int[key_int]), str):
                    flattened_dict_list_int.append(flatten_dict_to_list(dict_int[key_int]))
                else:
                    [flattened_dict_list_int.append(i) for i in flatten_dict_to_list(dict_int[key_int])]
    return flattened_dict_list_int